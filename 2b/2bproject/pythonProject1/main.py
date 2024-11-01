import re

import matplotlib
import numpy as np
import matplotlib.pyplot as plt
import scipy
import openpyxl
from scipy.optimize import curve_fit
from sklearn.linear_model import Ridge

matplotlib.use('TkAgg')
wb = openpyxl.Workbook()
sheet = wb.active

data_text = open('Task_2b.txt', 'r')

nx = int(re.findall(r'\d+', data_text.readline().strip())[0])
ny = int(re.findall(r'\d+', data_text.readline().strip())[0])

_ = data_text.readline().strip()

Y = []
x_points = []

for i in range(nx):
    line = list(map(float, re.findall(r'[-+]?(?:\d*\.*\d+)', data_text.readline().strip())))
    x_points.append(line[0])
    Y.append(line[1:])

data_text.close()

k = 2.535  # толерантный множитель
t = 2.31  # коэффициент стьюдента

i = 0
sall = []
avg_y = []
lcis = []
ucis = []
ltols = []
utols = []

for i in range(nx):
    avg_yi = np.mean(Y[i])
    avg_y.append(avg_yi)
    variance = 1 / (ny - 1) * sum([(yk - avg_yi) ** 2 for yk in Y[i]])
    sall.append(variance)
    lci = avg_yi - t * (variance ** (1 / 2)) / (ny ** (1 / 2))
    lcis.append(lci)
    uci = avg_yi + t * (variance ** (1 / 2)) / (ny ** (1 / 2))
    ucis.append(uci)
    ltol = avg_yi - (variance ** (1 / 2)) * k
    ltols.append(ltol)
    utol = avg_yi + (variance ** (1 / 2)) * k
    utols.append(utol)

    i += 1

kochren = max(sall) / sum(sall)

se2 = 1 / nx * sum(sall)  # средняя оценка дисперсии
q = 0  # начальная степень полинома
Fp = 1000  # начальное значение Fp критерия
while Fp > scipy.stats.f.ppf(0.95, nx - q - 1, ny - 1):
    X = np.vander(x_points, N=q + 1, increasing=True)  # матрица X
    a = np.linalg.inv(X.T @ X) @ X.T @ np.array(avg_y)  # оценка коэффициентов
    for i in range(len(a)):
        sheet.cell(row=1, column=i + 1, value=a[i])
    Sa = 1 / ny * se2 * (np.linalg.inv(X.T @ X))  # ковариационная матрица оценок коэффициентов
    print(np.linalg.cond(Sa))
    for i in range(len(Sa)):
        for j in range(len(Sa[i])):
            sheet.cell(row=i + 3, column=j + 1, value=Sa[i][j])
    Ra = np.corrcoef(Sa, rowvar=False)
    Fp = 1 / (nx - q - 1) * ny * se2 ** (-1) * (X @ a - np.array(avg_y)).T @ (X @ a - np.array(avg_y))
    for i in range(len(Ra)):
        for j in range(len(Ra[i])):
            sheet.cell(row=i + len(Sa), column=j + 1, value=Ra[i][j])
    wb.save('sa.xlsx')

    approximation = X @ a

    plt.scatter(x_points, avg_y, label='Средние точки', s=5)
    plt.plot(x_points, ltols, label='Нижний толерантный предел')
    plt.plot(x_points, utols, label='Верхний толерантный предел')
    plt.plot(x_points, lcis, label='Нижняя граница доверительного интервала')
    plt.plot(x_points, ucis, label='Верхняя граница доверительного интервала')
    plt.plot(x_points, approximation, label='Аппроксимация', lw=2.5)
    plt.legend()
    plt.show()

    q += 1

f = False
for i in range(ny):
    y = []
    for j in range(nx):
        y.append(Y[j][i])
    if not f:
        plt.scatter(x_points, y, color='red', s=2, label='Измерения')
        f = True
    else:
        plt.scatter(x_points, y, color='red', s=2)

plt.scatter(x_points, avg_y, label='Точки аппроксимации', s=7, color='black')

regress = scipy.stats.linregress(x_points, avg_y)
print(f'regress: y = {regress.slope} * x + ({regress.intercept}) ')
approximation = []
for x in x_points:
    approximation.append(regress.slope * x + regress.intercept)
plt.plot(x_points, approximation, label='regress', color='blue')

robust = scipy.stats.siegelslopes(avg_y, x_points)
print(f'robust: y = {robust.slope} * x + ({robust.intercept}) ')
approximation = []
for x in x_points:
    approximation.append(robust.slope * x + robust.intercept)
plt.plot(x_points, approximation, label='robust', color='orange')

polyfit = np.polynomial.polynomial.Polynomial.fit(x_points, avg_y, 1)
print(polyfit)
approximation = []
for x in x_points:
    approximation.append(polyfit(x))
plt.plot(x_points, approximation, label='polyfit', linestyle='dashed', dashes=(5, 5), color='red')

model_ridge = Ridge(alpha=1.0)
model_ridge.fit(np.array(x_points).reshape(-1, 1), avg_y)
approximation = model_ridge.predict(np.array(x_points).reshape(-1, 1))
print(f'ridge: y = {model_ridge.coef_[0]} * x + ({model_ridge.intercept_})')
plt.plot(x_points, approximation, label='ridge', color='green')

plt.legend()
plt.show()

polyfit = np.polynomial.polynomial.Polynomial.fit(x_points, avg_y, 40)
print(polyfit)
approximation = []
for x in x_points:
    approximation.append(polyfit(x))
plt.plot(x_points, approximation, label='polyfit')
plt.plot(x_points, lcis, label='Нижняя граница доверительного интервала')
plt.plot(x_points, ucis, label='Верхняя граница доверительного интервала')
plt.show()

x_smooth = np.linspace(-2, 2, 200)
approximation = np.interp(x_smooth, x_points, avg_y)
plt.plot(x_smooth, approximation)
plt.plot(x_points, lcis, label='Нижняя граница доверительного интервала')
plt.plot(x_points, ucis, label='Верхняя граница доверительного интервала')
plt.show()

approximation = scipy.interpolate.PchipInterpolator(np.array(x_points), np.array(avg_y))
x_smooth = np.linspace(-2, 2, 200)
plt.plot(x_smooth, approximation(x_smooth))
plt.plot(x_points, lcis, label='Нижняя граница доверительного интервала')
plt.plot(x_points, ucis, label='Верхняя граница доверительного интервала')
plt.show()

approximation = scipy.interpolate.CubicSpline(x_points, avg_y)
x_smooth = np.linspace(-2, 2, 200)
plt.plot(x_smooth, approximation(x_smooth))
plt.plot(x_points, lcis, label='Нижняя граница доверительного интервала')
plt.plot(x_points, ucis, label='Верхняя граница доверительного интервала')
plt.show()


def model_func(x, alpha, beta, a7, a6, a5, a4, a3, a2, a1, a0):
    return (np.sin(alpha * x) + beta) * (a7 * x**7 + a6 * x**6 + a5 * x**5 + a4 * x**4 + a3 * x**3 + a2 * x**2 + a1 * x + a0)


initial_guess = [1, 1, 1, 1, 1, 1, 1, 1, 1, 1]
params = curve_fit(model_func, x_points, avg_y, p0=initial_guess)[0]
print(params)
alpha, beta, a7, a6, a5, a4, a3, a2, a1, a0 = params
x_smooth = np.linspace(-2, 2, 200)
y = model_func(x_smooth, alpha, beta, a7, a6, a5, a4, a3, a2, a1, a0)
plt.plot(x_smooth, y)
plt.plot(x_points, lcis, label='Нижняя граница доверительного интервала')
plt.plot(x_points, ucis, label='Верхняя граница доверительного интервала')
plt.show()
